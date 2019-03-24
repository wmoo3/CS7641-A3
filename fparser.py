import numpy as np
import pandas as pd
import os
from openpyxl import load_workbook

fs = 'C:/Personal/GA Tech/CS7641/Assignment3'

subdirs = ["BASE", "PCA", "ICA", "RP", "RF"]
subdirs2 = ["PCA", "ICA", "RP", "RF"]
datasets1 = ["Diamonds", "Digits"]
datasets2 = ["diamonds", "digits"]
studies1 = ["logliklihood","SSE"]
studies2 = [" acc", " adjMI"]
studies3 = [" dim red"]
studies4 = [ " cluster GMM", " cluster Kmeans"]
studies5 = [[" variance"], [" recon"], [" coef", " recon"], [" importance"]]

'''
for dataset in datasets2:
    i=0
    for study in studies5:
        for s in study:
            print (subdirs2[i]+"/"+dataset+" scree"+s+".csv")
            df = pd.read_csv(fs+"/"+subdirs2[i]+"/"+dataset+" scree"+s+".csv")
            if i == 0:
                path = fs+"/RESULTS/"+ dataset + " scree.xlsx"
                df.to_excel(path, sheet_name=subdirs2[i] + " -" + s)
                book = load_workbook(path)
                writer = pd.ExcelWriter(path, engine = 'openpyxl')
                writer.book = book
            else:
                df.to_excel(writer,sheet_name=subdirs2[i] + " -" + s)
        i += 1
    writer.save()
    writer.close()
    

for dataset in datasets1:
    i=1
    for study in studies4:
        for subdir in subdirs:
            print (dataset+study+".csv")
            df = pd.read_csv(fs+"/"+subdir+"/"+dataset+study+".csv")
            if i == 1:
                path = fs+"/RESULTS/"+ dataset + " NN_cluster_reduction.xlsx"
                df.to_excel(path, sheet_name=subdir + " -" + study)
                book = load_workbook(path)
                writer = pd.ExcelWriter(path, engine = 'openpyxl')
                writer.book = book
            else:
                df.to_excel(writer,sheet_name=subdir + " -" + study)
            i += 1
    writer.save()
    writer.close()



for dataset in datasets2:
    for study in studies3:
        i=1
        for subdir in subdirs2:
            df = pd.read_csv(fs+"/"+subdir+"/"+dataset+study+".csv")
            print (dataset+study+".csv")
            if i == 1:
                path = fs+"/RESULTS/"+ dataset + " NN_dimension_reduction.xlsx"
                df.to_excel(path, sheet_name=subdir)
                book = load_workbook(path)
                writer = pd.ExcelWriter(path, engine = 'openpyxl')
                writer.book = book
            else:
                df.to_excel(writer,sheet_name=subdir)
            i += 1
    writer.save()
    writer.close()
    
'''
for dataset in datasets1:
    for study in studies2:
        i=1
        for subdir in subdirs:
            df = pd.read_csv(fs+"/"+subdir+"/run3/"+dataset+study+".csv")
            df['type'] = subdir
            print (study+".csv")
            if i == 1:
                df1 = df.copy()
            else:
                df1=df1.append(df)
            i += 1
        df1.to_csv(fs+"/RESULTS/"+dataset+study+".csv")
    
'''
for dataset in datasets2:
    i=1
    for subdir in subdirs:
        df = pd.read_csv(fs+"/"+subdir+"/"+dataset+"2D.csv")
        print (dataset+"2D.csv")
        if i == 1:
            df1 = df.copy()
        else:
            df1=df1.join(df,how='outer',rsuffix="_"+subdir)
        i += 1
    df1.to_csv(fs+"/RESULTS/"+dataset+"2D.csv")


for study in studies1:
    i=1
    for subdir in subdirs:
        df = pd.read_csv(fs+"/"+subdir+"/run3/"+study+".csv")
        print (study+".csv")
        if i == 1:
            df1 = df.copy()
        else:
            df1=df1.join(df,how='outer',rsuffix="_"+subdir)
        i += 1
    df1.to_csv(fs+"/RESULTS/"+study+".csv")
'''

